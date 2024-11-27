import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def compare_and_highlight(excel_file1, excel_file2, excel_file3, output_file):
    """比较三个 Excel 文件，并标记不同的 category 为黄色"""

    # 读取三个 Excel 文件的内容
    df1 = pd.read_excel(excel_file1, engine='openpyxl')
    df2 = pd.read_excel(excel_file2, engine='openpyxl')
    df3 = pd.read_excel(excel_file3, engine='openpyxl')

    # 载入 Excel 文件以便修改（用于高亮显示）
    wb1 = load_workbook(excel_file1)
    wb2 = load_workbook(excel_file2)
    wb3 = load_workbook(excel_file3)

    # 假设只需要处理第一个工作表
    ws1 = wb1.active
    ws2 = wb2.active
    ws3 = wb3.active

    # 使用黄色高亮颜色填充
    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 创建一个字典存储行信息：以 (conversation_id, session_id) 为键
    compare_dict = {}

    for df, ws in zip([df1, df2, df3], [ws1, ws2, ws3]):
        for index, row in df.iterrows():
            conversation_id = row["conversation_id"]
            session_id = row["session_id"]
            category = row["category"]

            # (conversation_id, session_id) 作为键存储
            if (conversation_id, session_id) not in compare_dict:
                compare_dict[(conversation_id, session_id)] = [None, None, None]

            if df is df1:
                compare_dict[(conversation_id, session_id)][0] = category
            elif df is df2:
                compare_dict[(conversation_id, session_id)][1] = category
            else:
                compare_dict[(conversation_id, session_id)][2] = category

    # 对比并标记差异
    for (conversation_id, session_id), categories in compare_dict.items():
        if len(set(categories)) > 1:  # 如果三个文件中的 category 不同
            # 找到每个文件中相应的行并标黄
            for i, (df, ws) in enumerate(zip([df1, df2, df3], [ws1, ws2, ws3])):
                row_index = df[(df["conversation_id"] == conversation_id) & (df["session_id"] == session_id)].index[
                                0] + 2  # Excel 的行是从 1 开始的
                if categories[i] != categories[0]:  # 如果 category 不一致，则标黄
                    if i == 0:
                        ws.cell(row=row_index, column=df.columns.get_loc("category") + 1).fill = highlight
                    elif i == 1:
                        ws.cell(row=row_index, column=df.columns.get_loc("category") + 1).fill = highlight
                    else:
                        ws.cell(row=row_index, column=df.columns.get_loc("category") + 1).fill = highlight

    # 保存修改后的文件
    wb1.save(output_file[0])
    wb2.save(output_file[1])
    wb3.save(output_file[2])

    print("Files have been compared and differences have been highlighted.")


def compare_excel_folders(excel_folder1, excel_folder2, excel_folder3, output_folder):
    """遍历文件夹中的 Excel 文件，进行比较"""

    # 获取文件夹中的所有 Excel 文件
    excel_files1 = [f for f in os.listdir(excel_folder1) if f.endswith('.xlsx')]

    # 遍历文件夹1中的每个 Excel 文件，查找文件夹2和3中同名的文件
    for excel_file in excel_files1:
        file1_path = os.path.join(excel_folder1, excel_file)
        file2_path = os.path.join(excel_folder2, excel_file)
        file3_path = os.path.join(excel_folder3, excel_file)

        # 确保三个文件都存在
        if os.path.exists(file2_path) and os.path.exists(file3_path):
            output_file = (
                os.path.join(output_folder, f"updated_{excel_file}"),
                os.path.join(output_folder, f"updated_{excel_file}"),
                os.path.join(output_folder, f"updated_{excel_file}")
            )
            compare_and_highlight(file1_path, file2_path, file3_path, output_file)
        else:
            print(f"Warning: Missing corresponding files for {excel_file}. Skipping.")


if __name__ == "__main__":
    # 输入的文件夹路径
    excel_folder1 = "E:/devwildchat/data/result-1"  # Excel 文件夹1路径
    excel_folder2 = "E:/devwildchat/data/result-2"  # Excel 文件夹2路径
    excel_folder3 = "E:/devwildchat/data/result-3"  # Excel 文件夹3路径

    # 输出文件夹路径（保存处理后的 Excel 文件）
    output_folder = "E:/devwildchat/data/result"  # 结果保存文件夹路径

    # 调用函数进行比较
    compare_excel_folders(excel_folder1, excel_folder2, excel_folder3, output_folder)
